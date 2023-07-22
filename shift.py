import pandas as pd#紹介の人が多すぎると人数不足になるかもしれない
import tkinter as tk#(特に各場所の枠が少ないときとか)
import tkinter.filedialog as fd
import openpyxl
import glob
import human_class
secondFlag=False

root=tk.Tk()#tkinter
root.geometry("700x550")
filename=""#予め作っておく変数
message=tk.Message(text="")
member=[]#参加者のリスト
member1=[]#⇓
member2=[]
member3=[]
member4=[]
place_df_list=[]
savefilename=""

def DataFrame(fpath):
    input_file=pd.ExcelFile(fpath)#Excelファイルの読み込み
    sheet_name=input_file.sheet_names#シート名のリストを作成
    people_df=input_file.parse(sheet_name[0])#参加者のデータフレームを作成
    setup_df=input_file.parse(sheet_name[1])#場所のデータフレームを作成
    first_df=input_file.parse(sheet_name[2])
    global secondFlag
    if len(sheet_name)==4:
        second_df=input_file.parse(sheet_name[3])
        secondFlag=True
        return people_df,setup_df,first_df,second_df,sheet_name
    elif len(sheet_name)==3:
        return people_df,setup_df,first_df,sheet_name

def openFile():#ファイル名とパスを取得
    fpath=fd.askopenfilename()
    global filename,message#global変数にして情報を入れるのが最適だと思う
    filename=fpath
    message.destroy()
    message=tk.Message(text="選択中のファイルは: "+filename)#表示するファイル名を更新
    message.pack()

def combine(memberlist):#紹介者を結合する関数
    count=0
    while count<(len(memberlist)):#ここ地味に工夫してる
        a=memberlist[count].species#紹介者ペアの結合
        if a=="紹介":
            memberlist[count-1].name=memberlist[count-1].name+"&"+memberlist[count].name
            memberlist[count-1].species=memberlist[count-1].species+"&"+memberlist[count].species
            memberlist[count-1].number+=1
            del memberlist[count]
        else:
            count+=1

def savename():
    global savefilename
    savefilename=txt.get()

btn1=tk.Button(text="①Excelファイルを選択",command=openFile)#tkinter
txt = tk.Entry()
btn2=tk.Button(text="③確定する",command=savename)
btn3=tk.Button(text="④Excelファイルを出力",command=root.quit)#仮でtkinterを閉じるコマンドを作っておく
lbl=tk.Label(text="②出力するファイルの名前を入力(.xlsxは不要)")
imageLabel=tk.Label()
btn1.pack()
lbl.pack()
txt.pack()
btn2.pack()
btn3.pack()
imageLabel.pack()
tk.mainloop()

savefilename=savefilename+".xlsx"

people_df=DataFrame(filename)[0]
for h in range(len(DataFrame(filename))-1):
    place_df_list.append(DataFrame(filename)[h+1])


if secondFlag==True:  #2日目があるかないかで分岐を作っちゃう（これはある方 (0)
    for i in range(len(people_df)):#people_dfからhumanクラスを作成
        student=human_class.human(people_df["氏名"][i],people_df["種別"][i],
        people_df["学年"][i],people_df["性別"][i],people_df["部活"][i],people_df["設営"][i],
        people_df["1日目"][i],people_df["2日目"][i])#枠数(student.number)は1で固定
        member.append(student)
        

    
    setup_member=[]#設営がoの人をsetup_memberに入れる
    counts=0
    while counts<(len(member)):
        if member[counts].setup=="o":
            setup_member.append(member[counts])
            counts+=1
        else:
            counts+=1

    combine(setup_member)
    
    setup_df=DataFrame(filename)[1]
    setup_place=[]
    for k in range(len(setup_df)):
        setup_place.append(human_class.place(setup_df["場所"][k],(setup_df["枠数"][k])))

    for l in range(len(setup_place)):#場所ごとに見て人を入れていく
        order=0
        while setup_place[l].number>0:#場所の枠数が０より大きいとき
            if setup_member[order].number<=setup_place[l].number:#人の持つ枠数が場所の残りの枠数以下のとき
                setup_place[l].member.append(setup_member[order])
                setup_place[l].number-=setup_member[order].number
                del setup_member[order]
            else:
                order+=1
        
    """
    for m in range(len(setup_place)):
        print(setup_place[m].name)
        for n in range(len(setup_place[m].member)):
            print(setup_place[m].member[n].name)
    """
    
    #ここsetupのところと同じ構造なんだけど関数にするのだるいからコピーして変数だけ変えてる
    
    for i in range(len(people_df)):#people_dfからhumanクラスを作成#member1を再び作る
        student=human_class.human(people_df["氏名"][i],people_df["種別"][i],
        people_df["学年"][i],people_df["性別"][i],people_df["部活"][i],people_df["設営"][i],
        people_df["1日目"][i],people_df["2日目"][i])#枠数(student.number)は1で固定
        member1.append(student)

    first_member=[]#設営がoの人をfirst_memberに入れる
    count=0
    while count<(len(member1)):
        if member1[count].first=="o":
            first_member.append(member1[count])
            count+=1
        else:
            count+=1

    combine(first_member)

    first_df=DataFrame(filename)[2]
    first_place=[]
    for o in range(len(first_df)):
        first_place.append(human_class.place(first_df["場所"][o],(first_df["枠数"][o])))

    for p in range(len(first_place)):#場所ごとに見て人を入れていく
        order=0
        while first_place[p].number>0:#場所の枠数が０より大きいとき
            if first_member[order].number<=first_place[p].number:#人の持つ枠数が場所の残りの枠数以下のとき
                first_place[p].member.append(first_member[order])
                first_place[p].number-=first_member[order].number
                del first_member[order]
            else:
                order+=1
    """
    for q in range(len(first_place)):
            print(first_place[q].name)
            for r in range(len(first_place[q].member)):
                print(first_place[q].member[r].name)
    """
    
    for i in range(len(people_df)):#people_dfからhumanクラスを作成#member2を再び作る
        student=human_class.human(people_df["氏名"][i],people_df["種別"][i],
        people_df["学年"][i],people_df["性別"][i],people_df["部活"][i],people_df["設営"][i],
        people_df["1日目"][i],people_df["2日目"][i])#枠数(student.number)は1で固定
        member2.append(student)

    second_member=[]#2日目がoの人をsecond_memberに入れる
    count=0
    while count<(len(member2)):
        if member2[count].second=="o":
            second_member.append(member2[count])
            count+=1
        else:
            count+=1

    combine(second_member)#second_member内で紹介者を結合
    
    second_df=DataFrame(filename)[3]
    second_place=[]
    for s in range(len(second_df)):
        second_place.append(human_class.place(second_df["場所"][s],(second_df["枠数"][s])))
    
    for t in range(len(second_place)):#場所ごとに見て人を入れていく
        order=0
        while second_place[t].number>0:#場所の枠数が０より大きいとき
            if second_member[order].number<=second_place[t].number:#人の持つ枠数が場所の残りの枠数以下のとき
                second_place[t].member.append(second_member[order])
                second_place[t].number-=second_member[order].number
                del second_member[order]
            else:
                order+=1
    """
    for u in range(len(second_place)):
            print(second_place[u].name)
            for v in range(len(second_place[u].member)):
                print(second_place[u].member[v].name)
    
    for w in range(len(second_member)):
        print(second_member[w].name)
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = DataFrame(filename)[4][1]
    wb.create_sheet(title=DataFrame(filename)[4][2])
    wb.create_sheet(title=DataFrame(filename)[4][3])

    write_ws = wb[DataFrame(filename)[4][1]]#setup
    write_ws1 = wb[DataFrame(filename)[4][2]]#firstdaay
    write_ws2 = wb[DataFrame(filename)[4][3]]#secondday

    write_ws.cell(row=1, column=1, value="氏名")
    write_ws.cell(row=1, column=2, value="種別")
    write_ws.cell(row=1, column=3, value="学年")

    write_ws1.cell(row=1, column=1, value="氏名")
    write_ws1.cell(row=1, column=2, value="種別")
    write_ws1.cell(row=1, column=3, value="学年")

    write_ws2.cell(row=1, column=1, value="氏名")
    write_ws2.cell(row=1, column=2, value="種別")
    write_ws2.cell(row=1, column=3, value="学年")
    
    count=2
    for x in range(len(setup_place)):
        write_ws.cell(row=count, column=1, value=setup_place[x].name)
        count+=1
        for y in range(len(setup_place[x].member)):
            write_ws.cell(row=count, column=1, value=setup_place[x].member[y].name)
            write_ws.cell(row=count, column=2, value=setup_place[x].member[y].species)
            write_ws.cell(row=count, column=3, value=setup_place[x].member[y].grade)
            count+=1

    count=2
    for z in range(len(first_place)):
        write_ws1.cell(row=count, column=1, value=first_place[z].name)
        count+=1
        for aa in range(len(first_place[z].member)):
            write_ws1.cell(row=count, column=1, value=first_place[z].member[aa].name)
            write_ws1.cell(row=count, column=2, value=first_place[z].member[aa].species)
            write_ws1.cell(row=count, column=3, value=first_place[z].member[aa].grade)
            count+=1
    
    count=2
    for bb in range(len(second_place)):
        write_ws2.cell(row=count, column=1, value=second_place[bb].name)
        count+=1
        for cc in range(len(second_place[bb].member)):
            write_ws2.cell(row=count, column=1, value=second_place[bb].member[cc].name)
            write_ws2.cell(row=count, column=2, value=second_place[bb].member[cc].species)
            write_ws2.cell(row=count, column=3, value=second_place[bb].member[cc].grade)
            count+=1
    
    wb.save(savefilename)#2日目ありの場合は完成(エクセルのセルサイズとかは工夫していいかも)
    #あとは1日目しかない場合も作る。オープンキャンパスバイトのときみたいに設営なしのパターンも簡単に付け足せると思う